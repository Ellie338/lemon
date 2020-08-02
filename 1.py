# encoding:utf-8
import requests
import openpyxl
import jsonpath
import os

# 读Excel中读测试用例
def read_data_excel(filename,sheetname):
    if os.path.exists(filename):     # os.path.exists()就是判断括号里的文件是否存在的意思，括号内的可以是文件路径。
        test_excel = openpyxl.load_workbook(filename)
        sheet = test_excel[sheetname]
        max_row = sheet.max_row
        list1 = []
        for i in range(2,max_row + 1,1):
            dict_1 = dict(
                case_id = sheet.cell(row = i,column = 1).value,
                interface = sheet.cell(row = i,column = 2).value,
                method = sheet.cell(row = i,column = 4).value,
                url = sheet.cell(row = i,column = 5).value,
                data = sheet.cell(row = i,column = 6).value,
                expected = sheet.cell(row = i,column =7).value)  # 从excel中取出URL、data、excepted 3列的内容，并强制转换成字典格式
            list1.append(dict_1)
        return list1
    else:
        return '文件不存在'


# 写入测试结果到excel
def write_result_excel(filename,sheetname,row,column,final_result):
    test_excel = openpyxl.load_workbook(filename)                      # 加载测试用例
    sheet = test_excel[sheetname]                                      # 取出sheet
    final_result1 = sheet.cell(row = row,column = column)     # 取出单元格内容   2行8列  .value == 取出值
    final_result1.value = final_result
    test_excel.save(filename)   # 保存测试用例                           # 行与列会发生变化 === 参数化  定义成形参
    return  final_result                              # 结果 = 执行结果和预期结果相比较，才能得到正确结果

# 发送请求
def send_request_no_token(method,url,data,header = {"X-Lemonban-Media-Type": "lemonban.v2","Content-Type": "application/json"}):
    if method.lower() == 'post':
        res1 = requests.post(url = url,json = data,headers = header).json()
    if method.lower() == 'get':
        res1 = requests.get(url = url,json = data,headers = header).json()
    if method.lower() == 'patch':
        res1 = requests.patch(url = url,json = data,headers = header).json()
    return res1

 # 执行接口
def exectue(filename,sheetname,log = None):
    test_case = read_data_excel(filename,sheetname)
    if not isinstance(test_case, str):
        for case in test_case:
            case_id = case['case_id']
            interface = case['interface']
            url = case['url']
            method = case['method']
            data = eval(case['data'])      # {member_id:'',}
            member_id = data.get('member_id')
            expected = eval(case['expected'])
            expected_msg = expected['msg']     # 预期结果的msg
            if interface == 'register' or interface == 'login':
                re_lo_rel_request = send_request_no_token(method,url,data)
            else:
                login_request = send_request_no_token(method = method,url = 'http://api.lemonban.com/futureloan/member/login',data = log)
                token = jsonpath.jsonpath(login_request,'$..token')[0]
                memberid = jsonpath.jsonpath(login_request,'$..id')[0]
                header_token = {"X-Lemonban-Media-Type": "lemonban.v2",
		                        "Content-Type": "application/json",
		                        "Authorization": "Bearer" + " " + token}
                if data is not None:
                    if interface == 'recharge' or interface == 'loan_add' or interface == 'withdraw' or interface == 'update':
                        if member_id is not None :
                            if member_id == 66666666:
                                real_request = send_request_no_token(method=method, url=url, data=data, header=header_token)
                            else:
                                data['member_id'] = memberid     # 将提取出来的memberid关联放在data中
                                real_request = send_request_no_token(method = method,url = url,data = data,header = header_token)
                        else:
                            real_request = send_request_no_token(method=method, url=url, data=data, header=header_token)
                    else:
                        real_request = send_request_no_token(method = method,url = url,data = data)    # 翻页
                else:
                    if url == 'http://120.78.128.25:8766/futureloan/member/{}/info':
                        N_url = url.format(memberid)      # 新的url包含memberid
                        real_request = send_request_no_token(method,data = None,url = N_url)      # 正常访问id信息
                    else:
                        real_request = send_request_no_token(method,data = None,url = url)
            real_msg = real_request['msg']
            print('*' * 30)
            print('预期结果为{}'.format(expected_msg))
            print('实际结果为{}'.format(real_msg))

            if real_msg == expected_msg:
                print('第{}条用例通过'.format(case_id))
                final_result = 'Pass'
                s = write_result_excel(filename, sheetname, case_id + 1, 8, final_result)
            else:
                print('第{}条用例不通过'.format(case_id))
                final_result = 'Fail'
                s = write_result_excel(filename, sheetname, case_id + 1, 8, final_result)
        return s
    else:
        return '文件不存在'

# print(exectue('test_case_api.xlsx','register'))  # 注册
# print(exectue('test_case_api.xlsx','login'))     # 登录

# 普通用户
login_body = {
    "mobile_phone":"15512265678",
    "pwd": "12345678"
}

# 管理员账号
login0_body = {
    "mobile_phone": "15815541666",
    "pwd": "lemon123456"
}

print(exectue('test_case_api.xlsx', 'recharge',login_body))   # 充值
# print(exectue('test_case_api.xlsx', 'user_info',login_body))  # 获取用户信息
# print(exectue('test_case_api.xlsx', 'update',login_body))     # 更新昵称
# print(exectue('test_case_api.xlsx', 'loans',login_body))      # 获取项目列表

# ---------------还没有做-----------------------------------------------
# print(exectue('test_case_api.xlsx', 'loan_add',login_body))  # 加标
# print(exectue('test_case_api.xlsx', 'loan_audit',login0_body)) # 审核
#投资






